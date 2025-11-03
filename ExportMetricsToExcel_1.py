from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.model_selection import cross_val_predict
from sklearn.metrics import precision_score, recall_score, f1_score
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import pandas as pd
import time
import os
from pathlib import Path

class ExportCalculatedMetrics:
    '''Entrada:
        Recebe uma lista de métricas (metrics_list), onde cada elemento é um dicionário contendo resultados de modelos de ML (ex.: precisão, recall, F1-score).
    Saída:
        Gera um arquivo Excel (testeeee.xlsx) com tabelas formatadas comparando o desempenho dos modelos. O arquivo é salvo em EXPORT_TEST/ e aberto automaticamente.
    Funcionalidade:
        Organiza e formata métricas de modelos (como LogisticRegression, RandomForest, etc.) em uma planilha legível.
    '''

    def __init__(self, metrics_list, path_name):
        self.metrics_list = metrics_list
        self.path_name = path_name

    def format_table(self, ws, title, num_columns):
        '''!! FALTA DESCRIÇÃO !!'''
        # Adicionar título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col in range(1, num_columns + 1):
            cell = cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Formatar células de dados
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=num_columns):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
        
        # Ajustar largura das colunas (ignorando células mescladas)
        for col_idx in range(1, num_columns + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                try:
                    if cell.value and not (hasattr(cell, 'merged') and cell.merged):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    def add_tabel(self, ws, tabela_dict, linha_inicio, coluna_inicio):
        tabela = tabela_dict.copy()
        title = tabela.pop('title')
        df = pd.DataFrame(tabela)
        num_colunas = len(df.columns)
        
        # Mesclar células para o título (4 colunas)
        ws.merge_cells(start_row=linha_inicio, start_column=coluna_inicio,
                    end_row=linha_inicio, end_column=coluna_inicio + num_colunas - 1)
        
        # Adicionar título centralizado
        cell = ws.cell(row=linha_inicio, column=coluna_inicio, value=title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adicionar cabeçalhos
        for col_num, column in enumerate(df.columns, start=coluna_inicio):
            cell = ws.cell(row=linha_inicio+1, column=col_num, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Adicionar dados
        for row_num, row in df.iterrows():
            for col_num, value in enumerate(row, start=coluna_inicio):
                cell = ws.cell(row=row_num+linha_inicio+2, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
        
        return num_colunas, len(df) + 2  # +2 para título e cabeçalho

    def export_metrics(self):

        # Configurações
        COLUNAS_POR_LINHA = 2
        ESPACO_ENTRE_TABELAS = 2
        LINHA_INICIAL = 1
        COLUNA_INICIAL = 1

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparação de Modelos"

        # Posicionar tabelas
        current_row = LINHA_INICIAL
        current_col = COLUNA_INICIAL
        max_linhas_por_linha = 0
        model_columns = []  # Armazenar índices das colunas 'model'

        for i, tabela in enumerate(self.metrics_list):
            largura, altura = self.add_tabel(ws, tabela, current_row, current_col)
            model_columns.append(current_col)  # Adicionar índice da coluna 'model'
            max_linhas_por_linha = max(max_linhas_por_linha, altura)
            
            if (i + 1) % COLUNAS_POR_LINHA == 0:
                current_row += max_linhas_por_linha + 1
                current_col = COLUNA_INICIAL
                max_linhas_por_linha = 0
            else:
                current_col += largura + ESPACO_ENTRE_TABELAS

        # Ajustar largura das colunas, exceto para as colunas 'model'
        for col in ws.columns:
            col_idx = col[0].column
            if col_idx in model_columns:  # Pular colunas 'model'
                continue
            max_length = 0
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        path_name_ = self.path_name + ".xlsx"

        wb.save(path_name_)
        print("Nome do ficheiro:", path_name_)
        print("Arquivo exportado com sucesso!")

        # ABRIR AUTOMATICAMENTE O FICHEIRO
        os.startfile(Path(path_name_))